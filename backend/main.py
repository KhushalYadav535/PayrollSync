from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from pydantic import BaseModel
import logging

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(levelname)s │ %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(title="PayrollSync ECR API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
OUTPUT_DIR = "output"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Pydantic models ───────────────────────────────────────────────────────────
class ProcessingResult(BaseModel):
    success: bool
    message: str
    files_generated: List[str]
    errors: List[str]
    summary: Optional[Dict[str, Any]] = None

class FileInfo(BaseModel):
    filename: str
    type: str

class DownloadResponse(BaseModel):
    files: List[FileInfo]

# ── Constants ─────────────────────────────────────────────────────────────────
EPS_WAGE_CAP   = 15000
EDLI_WAGE_CAP  = 15000
EE_RATE        = 0.12
ER_EPS_FACTOR  = 8.33 / 12          # (8.33/12) × EE Share → ER Share

# ── Header-detection keywords ─────────────────────────────────────────────────
UAN_KEYS   = {"uan", "pf no", "pf number", "account no", "pf"}
NAME_KEYS  = {"employee name", "member name", "emp name", "name", "employee"}
WAGE_KEYS  = {"gross salary", "gross wages", "gross", "salary", "wages", "basic"}
PF_KEYS    = {"pf deducted", "employee pf", "epf", "pf", "deduction"}

SKIP_VALUES = {
    "total", "grand total", "summary", "subtotal", "sub-total",
    "net total", "note", "notes", ""
}

# ── Utility helpers ────────────────────────────────────────────────────────────
def _norm(val) -> str:
    """Lowercase, trim, collapse whitespace, strip hidden chars."""
    if pd.isna(val):
        return ""
    s = str(val)
    s = s.encode("ascii", "ignore").decode()     # drop non-ASCII
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"[^\x20-\x7E]", "", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    # remove special chars except letters, digits, space
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def clean_uan(val) -> str:
    if pd.isna(val):
        return ""
    s = str(val).strip()
    # Handle scientific notation (e.g. 1.0101e+11)
    if re.search(r"[eE]", s):
        try:
            s = str(int(float(s)))
        except Exception:
            pass
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"[^0-9]", "", s)
    return s

def clean_name(val) -> str:
    if pd.isna(val):
        return ""
    s = str(val).strip().upper()
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def to_numeric_safe(val) -> Optional[float]:
    if pd.isna(val):
        return None
    s = str(val).replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "").strip()
    try:
        return float(s)
    except Exception:
        return None

def is_skip_row(raw_name, raw_wage) -> bool:
    n = _norm(raw_name)
    w = _norm(raw_wage)
    if n in SKIP_VALUES or w in SKIP_VALUES:
        return True
    # Location / header words appearing as data
    if any(kw in n for kw in ["total", "grand", "summary", "note", "sl no", "sr no"]):
        return True
    return False

# ── Header detection ──────────────────────────────────────────────────────────
def _score_row(row) -> int:
    all_kws = UAN_KEYS | NAME_KEYS | WAGE_KEYS | PF_KEYS
    score = 0
    for cell in row:
        n = _norm(cell)
        for kw in all_kws:
            if kw in n:
                score += 1
                break
    return score

def find_header_row(df_raw: pd.DataFrame) -> int:
    best_idx, best_score = 0, 0
    for idx, row in df_raw.head(50).iterrows():
        s = _score_row(row)
        if s > best_score:
            best_score = s
            best_idx = idx
    return best_idx

# ── Column mapping ────────────────────────────────────────────────────────────
def _find_col(cols_norm: list, keys: set, exclude=("amount", "share", "cont", "rate")):
    for kw in sorted(keys, key=len, reverse=True):   # longest match first
        for i, c in enumerate(cols_norm):
            if kw in c and not any(ex in c for ex in exclude):
                return i
    return None

def map_columns(df: pd.DataFrame):
    cols_norm = [_norm(c) for c in df.columns]
    ui = _find_col(cols_norm, UAN_KEYS, exclude=("amount",))
    ni = _find_col(cols_norm, NAME_KEYS, exclude=())
    wi = _find_col(cols_norm, WAGE_KEYS, exclude=("pf", "eps", "epf", "er", "ee", "cont"))
    pi = _find_col(cols_norm, PF_KEYS, exclude=("wage",))
    uan_col  = df.columns[ui] if ui is not None else None
    name_col = df.columns[ni] if ni is not None else None
    wage_col = df.columns[wi] if wi is not None else None
    pf_col   = df.columns[pi] if pi is not None and pi != wi else None
    return uan_col, name_col, wage_col, pf_col

# ── EPFO contribution logic ───────────────────────────────────────────────────
def calc_contributions(gross: float) -> Dict[str, int]:
    epf_wages  = gross                              # no cap on EPF wages
    eps_wages  = min(gross, EPS_WAGE_CAP)
    edli_wages = min(gross, EDLI_WAGE_CAP)
    ee_share   = round(EE_RATE * epf_wages)
    er_share   = round(ER_EPS_FACTOR * ee_share)            # ER Share = (8.33/12) × EE Share
    eps_cont   = ee_share - er_share                        # EPF portion retained
    return {
        "Gross Wages":       int(round(gross)),
        "EPF Wages":         int(round(epf_wages)),
        "EPS Wages":         int(round(eps_wages)),
        "EDLI Wages":        int(round(edli_wages)),
        "EE Share":          int(ee_share),
        "ER Share":          int(er_share),
        "EPS Contribution":  int(eps_cont),
        "NCP Days":          0,
        "Refund":            0,
    }

ECR_COLS = [
    "UAN", "Member Name", "Gross Wages", "EPF Wages", "EPS Wages",
    "EDLI Wages", "EE Share", "ER Share", "EPS Contribution", "NCP Days", "Refund"
]

# ── Month label from filename / sheet ────────────────────────────────────────
MONTH_MAP = {
    "jan": "JAN", "feb": "FEB", "mar": "MAR", "apr": "APR",
    "may": "MAY", "jun": "JUN", "jul": "JUL", "aug": "AUG",
    "sep": "SEP", "oct": "OCT", "nov": "NOV", "dec": "DEC",
}

def extract_month_label(source: str) -> str:
    s = source.upper()
    for k, v in MONTH_MAP.items():
        if k.upper() in s:
            # Try to get year too
            year_match = re.search(r"(20\d{2})", s)
            year = year_match.group(1)[2:] if year_match else datetime.now().strftime("%y")
            return f"{v}{year}"
    return datetime.now().strftime("%b%Y").upper()

# ── Excel audit workbook ──────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="EEF2FF")
BORDER_THIN  = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
WARN_FILL    = PatternFill("solid", fgColor="FEF3C7")
ERR_FILL     = PatternFill("solid", fgColor="FEE2E2")

def _style_header_row(ws, row_num: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

def _style_data_row(ws, row_num: int, ncols: int, alt: bool = False, fill=None):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill
        elif alt:
            cell.fill = ALT_FILL

def _autofit(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def write_audit_excel(
    ecr_df: pd.DataFrame,
    rejected: List[Dict],
    month_label: str,
    out_path: str,
):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Final ECR ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Final ECR"
    ws1.row_dimensions[1].height = 30
    for ci, col in enumerate(ECR_COLS, 1):
        ws1.cell(row=1, column=ci, value=col)
    _style_header_row(ws1, 1, len(ECR_COLS))

    for ri, (_, row) in enumerate(ecr_df[ECR_COLS].iterrows(), 2):
        for ci, col in enumerate(ECR_COLS, 1):
            ws1.cell(row=ri, column=ci, value=row[col])
        _style_data_row(ws1, ri, len(ECR_COLS), alt=(ri % 2 == 0))

    # Totals row
    tr = len(ecr_df) + 2
    ws1.cell(row=tr, column=1, value="TOTAL")
    ws1.cell(row=tr, column=2, value=f"{len(ecr_df)} Employees")
    numeric_cols = ["Gross Wages","EPF Wages","EPS Wages","EDLI Wages","EE Share","ER Share","EPS Contribution"]
    for ci, col in enumerate(ECR_COLS, 1):
        if col in numeric_cols:
            ws1.cell(row=tr, column=ci, value=int(ecr_df[col].sum()))
    _style_header_row(ws1, tr, len(ECR_COLS))
    ws1.cell(row=tr, column=1).font = Font(bold=True, color="FFD700", size=11)

    _autofit(ws1)
    _freeze(ws1)

    # ── Sheet 2: Validation Report ────────────────────────────────────────────
    ws2 = wb.create_sheet("Validation Report")
    v_cols = ["Row Ref", "Raw UAN", "Raw Name", "Raw Salary", "Rejection Reason"]
    ws2.row_dimensions[1].height = 30
    for ci, col in enumerate(v_cols, 1):
        ws2.cell(row=1, column=ci, value=col)
    _style_header_row(ws2, 1, len(v_cols))

    if rejected:
        for ri, rec in enumerate(rejected, 2):
            reason = rec.get("reason", "")
            for ci, col in enumerate(v_cols, 1):
                ws2.cell(row=ri, column=ci, value=rec.get(col, ""))
            fill = ERR_FILL if "missing" in reason.lower() or "invalid" in reason.lower() else WARN_FILL
            _style_data_row(ws2, ri, len(v_cols), fill=fill)
    else:
        ws2.cell(row=2, column=1, value="✅ No rejected records — all rows passed validation.")
        ws2.cell(row=2, column=1).font = Font(bold=True, color="16A34A")

    _autofit(ws2)
    _freeze(ws2)

    # ── Sheet 3: Summary ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    summary_rows = [
        ("Month", month_label),
        ("Total Employees", len(ecr_df)),
        ("Total Gross Wages", int(ecr_df["Gross Wages"].sum())),
        ("Total EPF Wages", int(ecr_df["EPF Wages"].sum())),
        ("Total EE Share", int(ecr_df["EE Share"].sum())),
        ("Total ER Share", int(ecr_df["ER Share"].sum())),
        ("Total EPS Contribution", int(ecr_df["EPS Contribution"].sum())),
        ("Total Rejected Rows", len(rejected)),
        ("Generated On", datetime.now().strftime("%d-%b-%Y %H:%M")),
    ]
    ws3.row_dimensions[1].height = 35
    ws3.cell(row=1, column=1, value=f"ECR SUMMARY — {month_label}")
    ws3.merge_cells("A1:B1")
    ws3.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF", size=13)
    ws3.cell(row=1, column=1).fill = HEADER_FILL
    ws3.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for ri, (label, value) in enumerate(summary_rows, 2):
        ws3.cell(row=ri, column=1, value=label).font = Font(bold=True, size=11)
        ws3.cell(row=ri, column=2, value=value).font = Font(size=11)
        ws3.cell(row=ri, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws3.cell(row=ri, column=2).alignment = Alignment(horizontal="right", vertical="center")
        if ri % 2 == 0:
            for c in [1, 2]:
                ws3.cell(row=ri, column=c).fill = ALT_FILL
        for c in [1, 2]:
            ws3.cell(row=ri, column=c).border = BORDER_THIN
        ws3.row_dimensions[ri].height = 22

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 25

    wb.save(out_path)

# ── Core processing engine ────────────────────────────────────────────────────
def process_excel_file(file_path: str, original_filename: str) -> ProcessingResult:
    errors: List[str] = []
    files_generated: List[str] = []

    # Wipe old outputs
    for f in os.listdir(OUTPUT_DIR):
        fp = os.path.join(OUTPUT_DIR, f)
        if os.path.isfile(fp):
            try:
                os.remove(fp)
            except Exception as e:
                logger.warning(f"Could not remove {f}: {e}")

    try:
        xls = pd.ExcelFile(file_path)
        all_ecr: List[Dict] = []
        all_rejected: List[Dict] = []

        base_name = os.path.splitext(original_filename)[0]
        month_label = extract_month_label(base_name)

        logger.info(f"Sheets found: {xls.sheet_names} | Month: {month_label}")

        for sheet_name in xls.sheet_names:
            try:
                df_raw = xls.parse(sheet_name, header=None)
                if df_raw.empty:
                    continue

                header_idx = find_header_row(df_raw)
                df = xls.parse(sheet_name, header=header_idx)

                if df.empty or len(df.columns) < 2:
                    errors.append(f"Sheet '{sheet_name}': Too few columns after header detection.")
                    continue

                uan_col, name_col, wage_col, pf_col = map_columns(df)

                if uan_col is None or name_col is None or wage_col is None:
                    errors.append(
                        f"Sheet '{sheet_name}': Could not detect required columns "
                        f"(UAN={uan_col}, Name={name_col}, Wage={wage_col})."
                    )
                    continue

                seen_keys = set()

                for raw_idx, row in df.iterrows():
                    raw_uan   = row[uan_col]
                    raw_name  = row[name_col]
                    raw_wage  = row[wage_col]
                    row_ref   = f"{sheet_name}:R{header_idx + raw_idx + 2}"

                    # Skip total / note rows
                    if is_skip_row(raw_name, raw_wage):
                        continue

                    uan   = clean_uan(raw_uan)
                    name  = clean_name(raw_name)
                    wage  = to_numeric_safe(raw_wage)

                    def _reject(reason):
                        all_rejected.append({
                            "Row Ref": row_ref,
                            "Raw UAN": str(raw_uan),
                            "Raw Name": str(raw_name),
                            "Raw Salary": str(raw_wage),
                            "Rejection Reason": reason,
                        })

                    # Validation
                    if not uan:
                        _reject("Missing or invalid UAN")
                        continue
                    if not re.fullmatch(r"\d+", uan):
                        _reject(f"UAN contains non-numeric chars: {uan}")
                        continue
                    if not name or len(name) < 3:
                        _reject("Missing or too-short employee name")
                        continue
                    if name in {"NAME", "EMP", "NAM", "EMPLOYEE"}:
                        _reject("Name appears to be a header/placeholder")
                        continue
                    if wage is None or wage <= 0:
                        _reject(f"Invalid salary: {raw_wage}")
                        continue

                    # Deduplication on UAN + Salary
                    dedup_key = (uan, int(round(wage)))
                    if dedup_key in seen_keys:
                        _reject("Duplicate record (UAN + Salary)")
                        continue
                    seen_keys.add(dedup_key)

                    contrib = calc_contributions(wage)
                    row_data = {"UAN": uan, "Member Name": name, **contrib}
                    all_ecr.append(row_data)

                logger.info(f"Sheet '{sheet_name}': {len([r for r in all_ecr])} accepted so far")

            except Exception as e:
                errors.append(f"Sheet '{sheet_name}' error: {str(e)}")
                logger.exception(f"Error in sheet {sheet_name}")
                continue

        if not all_ecr:
            return ProcessingResult(
                success=False,
                message="No valid records found in any sheet.",
                files_generated=[],
                errors=errors,
            )

        ecr_df = pd.DataFrame(all_ecr)[ECR_COLS]
        ecr_df = ecr_df.reset_index(drop=True)

        # ── CSV ───────────────────────────────────────────────────────────────
        csv_name = f"ECR_UPLOAD_{month_label}.csv"
        csv_path = os.path.join(OUTPUT_DIR, csv_name)
        ecr_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        files_generated.append(csv_name)

        # ── TXT (#~# separated, no header) ───────────────────────────────────
        txt_name = f"ECR_UPLOAD_{month_label}.txt"
        txt_path = os.path.join(OUTPUT_DIR, txt_name)
        with open(txt_path, "w", encoding="utf-8") as tf:
            for _, row in ecr_df.iterrows():
                line = "#~#".join(str(int(v)) if isinstance(v, float) else str(v) for v in row)
                tf.write(line + "\n")
        files_generated.append(txt_name)

        # ── Audit Excel ───────────────────────────────────────────────────────
        audit_name = f"ECR_AUDIT_{month_label}.xlsx"
        audit_path = os.path.join(OUTPUT_DIR, audit_name)
        write_audit_excel(ecr_df, all_rejected, month_label, audit_path)
        files_generated.append(audit_name)

        summary = {
            "total_employees":    len(ecr_df),
            "total_gross":        int(ecr_df["Gross Wages"].sum()),
            "total_epf_wages":    int(ecr_df["EPF Wages"].sum()),
            "total_ee_share":     int(ecr_df["EE Share"].sum()),
            "total_er_share":     int(ecr_df["ER Share"].sum()),
            "total_eps_cont":     int(ecr_df["EPS Contribution"].sum()),
            "rejected_rows":      len(all_rejected),
            "month":              month_label,
        }

        logger.info(f"Generated {len(files_generated)} files for {month_label}")

        return ProcessingResult(
            success=True,
            message=(
                f"Successfully processed {len(ecr_df)} employees for {month_label}. "
                f"{len(all_rejected)} row(s) rejected."
            ),
            files_generated=files_generated,
            errors=errors,
            summary=summary,
        )

    except Exception as e:
        logger.exception("Fatal processing error")
        return ProcessingResult(
            success=False,
            message=f"Fatal error: {str(e)}",
            files_generated=[],
            errors=[str(e)],
        )

# ── FastAPI routes ─────────────────────────────────────────────────────────────
@app.post("/upload", response_model=ProcessingResult)
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx / .xls) are accepted.")

    file_path = os.path.join(UPLOAD_DIR, file.filename)
    try:
        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)
        logger.info(f"Uploaded: {file.filename} ({len(content)/1024:.1f} KB)")
        return process_excel_file(file_path, file.filename)
    except Exception as e:
        logger.exception("Upload route error")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/downloads", response_model=DownloadResponse)
async def get_downloads():
    files = []
    for fname in os.listdir(OUTPUT_DIR):
        fpath = os.path.join(OUTPUT_DIR, fname)
        if os.path.isfile(fpath):
            ext = fname.rsplit(".", 1)[-1].lower()
            ftype = {"csv": "CSV", "txt": "TEXT", "xlsx": "Excel"}.get(ext, "File")
            files.append(FileInfo(filename=fname, type=ftype))
    return DownloadResponse(files=files)

@app.get("/download/{filename}")
async def download_file(filename: str):
    # Sanitise path
    filename = os.path.basename(filename)
    file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found.")
    return FileResponse(file_path, media_type="application/octet-stream", filename=filename)

@app.get("/")
async def root():
    return {"message": "PayrollSync ECR API v2.0", "status": "running"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
